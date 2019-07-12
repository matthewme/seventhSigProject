#!/usr/bin/env python
# Author: Matthew Martinez
# Contact: mattmartinez3n@gmail.com
# Date: 2019

import os
import xlrd
import sqlite3
import zipfile
from peewee import *
from openpyxl import *
import filetype
import shutil

# Connect to the SQLite database
db = SqliteDatabase('D:/Program Files/SQLiteStudio/convergenceDB', pragmas={'foreign_keys': 1})


class Installation(Model):
    id = AutoField()
    installation_name = CharField()

    class Meta:
        # data is coming from matthewDB.db
        database = db


class G1(Model):
    # These are all the fields it has
    # match up CharField/IntegerField/etc with correct type
    id = AutoField()
    date = DateField()
    major_command = CharField()
    unit = CharField()
    installation = CharField()
    address = CharField()
    city = CharField()
    state = CharField()
    poc = CharField()
    poc_email = CharField()
    poc_phone = CharField()
    team_member_name = CharField()
    team_member_phone = CharField()
    team_member_email = CharField()
    team_member_signature = CharField()
    num_supported_units = CharField()
    num_of_personnel = CharField()
    union_reps = CharField()
    g1_fk_id = IntegerField()

    class Meta:
        # data is coming from convergenceDB.db
        database = db


class G2(Model):
    # These are all the fields it has
    # match up CharField/IntegerField/etc with correct type
    id = AutoField()
    date = DateField()
    organization = CharField()
    unit = CharField()
    address = CharField()
    point_of_contact = CharField()
    email = CharField()
    phone_number = CharField()
    name = CharField()
    second_phone_number = CharField()
    second_email = CharField()
    signature = CharField()
    detailed_data_a = CharField()
    detailed_data_b = CharField()
    detailed_data_c = CharField()
    detailed_data_d = CharField()
    detailed_data_e = CharField()
    detailed_data_f = CharField()
    detailed_data_g = CharField()
    detailed_data_h = CharField()
    g2_fk_id = IntegerField()

    class Meta:
        # data is coming from convergenceDB.db
        database = db


class G3(Model):
    # These are all the fields it has
    # match up CharField/IntegerField/etc with correct type
    id = AutoField()
    date = DateField()
    organization = CharField()
    unit = CharField()
    address = CharField()
    point_of_contact = CharField()
    email = CharField()
    phone_number = CharField()
    name = CharField()
    second_phone_number = CharField()
    second_email = CharField()
    signature = CharField()
    detailed_data_a = CharField()
    detailed_data_b = CharField()
    detailed_data_c = CharField()
    detailed_data_d = CharField()
    detailed_data_e = CharField()
    detailed_data_f1 = CharField()
    detailed_data_f2 = CharField()
    detailed_data_f3 = CharField()
    detailed_data_f4 = CharField()
    detailed_data_f5 = CharField()
    detailed_data_f6 = CharField()
    g3_fk_id = IntegerField()

    class Meta:
        # data is coming from convergenceDB.db
        database = db


class G4C(Model):
    # These are all the fields it has
    # match up CharField/IntegerField/etc with correct type
    id = AutoField()
    date = DateField()
    organization = CharField()
    unit = CharField()
    address = CharField()
    point_of_contact = CharField()
    email = CharField()
    phone_number = CharField()
    name = CharField()
    second_phone_number = CharField()
    second_email = CharField()
    signature = CharField()
    detailed_data_a = CharField()
    detailed_data_b1 = CharField()
    detailed_data_b2 = CharField()
    detailed_data_b3 = CharField()
    detailed_data_b4 = CharField()
    detailed_data_b5 = CharField()
    detailed_data_b6 = CharField()
    g4c_fk_id = IntegerField()

    class Meta:
        # data is coming from convergenceDB.db
        database = db


class G4PF(Model):
    # These are all the fields it has
    # match up CharField/IntegerField/etc with correct type
    id = AutoField()
    date = DateField()
    organization = CharField()
    unit = CharField()
    address = CharField()
    point_of_contact = CharField()
    email = CharField()
    phone_number = CharField()
    name = CharField()
    second_phone_number = CharField()
    second_email = CharField()
    signature = CharField()
    detailed_data_a1 = CharField()
    detailed_data_a2 = CharField()
    detailed_data_a3 = CharField()
    detailed_data_a4 = CharField()
    detailed_data_a5 = CharField()
    detailed_data_b1 = CharField()
    detailed_data_b2 = CharField()
    detailed_data_b3 = CharField()
    detailed_data_b4 = CharField()
    g4pf_fk_id = IntegerField()

    class Meta:
        # data is coming from convergenceDB.db
        database = db


class G5(Model):
    # These are all the fields it has
    # match up CharField/IntegerField/etc with correct type
    id = AutoField()
    date = DateField()
    organization = CharField()
    unit = CharField()
    address = CharField()
    point_of_contact = CharField()
    email = CharField()
    phone_number = CharField()
    name = CharField()
    second_phone_number = CharField()
    second_email = CharField()
    signature = CharField()
    detailed_data_a = CharField()
    detailed_data_b = CharField()
    detailed_data_c = CharField()
    g5_fk_id = IntegerField()

    class Meta:
        # data is coming from convergenceDB.db
        database = db


class G6(Model):
    # These are all the fields it has
    # match up CharField/IntegerField/etc with correct type
    id = AutoField()
    date = DateField()
    organization = CharField()
    unit = CharField()
    address = CharField()
    point_of_contact = CharField()
    email = CharField()
    phone_number = CharField()
    name = CharField()
    second_phone_number = CharField()
    second_email = CharField()
    signature = CharField()
    detailed_data_a = CharField()
    detailed_data_b = CharField()
    detailed_data_c = CharField()
    g6_fk_id = IntegerField()

    class Meta:
        # data is coming from convergenceDB.db
        database = db


class G8(Model):
    # These are all the fields it has
    # match up CharField/IntegerField/etc with correct type
    id = AutoField()
    date = DateField()
    organization = CharField()
    unit = CharField()
    address = CharField()
    point_of_contact = CharField()
    email = CharField()
    phone_number = CharField()
    name = CharField()
    second_phone_number = CharField()
    second_email = CharField()
    signature = CharField()
    detailed_data_a = CharField()
    detailed_data_a1 = CharField()
    detailed_data_a2 = CharField()
    detailed_data_a3 = CharField()
    detailed_data_a4 = CharField()
    detailed_data_b1 = CharField()
    detailed_data_b2 = CharField()
    detailed_data_c1 = CharField()
    detailed_data_c2 = CharField()
    g8_fk_id = IntegerField()

    class Meta:
        # data is coming from convergenceDB.db
        database = db


class Admin(Model):
    admin1 = CharField()
    admin2 = CharField()
    admin3 = CharField()
    admin4 = CharField()
    admin5 = CharField()
    customer1 = CharField()
    customer2 = CharField()
    customer3 = CharField()
    customer4 = CharField()
    customer5 = CharField()
    customer6 = CharField()
    customer7 = CharField()
    name1 = CharField()
    name2 = CharField()
    name3 = CharField()
    name4 = CharField()
    name5 = CharField()
    name6 = CharField()
    phone1 = CharField()
    phone2 = CharField()
    phone3 = CharField()
    phone4 = CharField()
    phone5 = CharField()
    phone6 = CharField()
    email1 = CharField()
    email2 = CharField()
    email3 = CharField()
    email4 = CharField()
    email5 = CharField()
    email6 = CharField()
    tenants1 = CharField()
    tenants2 = CharField()
    tenants3 = CharField()
    tenants4 = CharField()
    tenants5 = CharField()
    tenants6 = CharField()
    tenants7 = CharField()
    admin_fk_id = IntegerField()

    class Meta:
        # data is coming from convergenceDB.db
        database = db


class PDFStorage(Model):
    # These are all the fields it has
    # match up CharField/IntegerField/etc with correct type
    id = AutoField()
    pdf_byte_data = BlobField()
    pdf_storage_fk_id = IntegerField()

    class Meta:
        # data is coming from convergenceDB.db
        database = db


# for i in range(len(wb.sheet_names())):# of sheets in workbook
def main():
    # Connect to DB
    db.connect()

    for root, dirs, files in os.walk('C:/Users/mattm/Desktop/spreadsheets'):
        xlsfiles = [_ for _ in files if _.endswith('.xlsm')]
        for xlsfile in xlsfiles:
            # wbXLRD = xlrd.open_workbook(os.path.join(root, xlsfile), on_demand=True)
            wbOPENPYXL = load_workbook(os.path.join(root, xlsfile))

            print(wbOPENPYXL.sheetnames)

            # ------------------Extract PDFs---------------------------------------
            filename = os.path.join(root, xlsfile)
            # print(filename)

            prefix = filename.split('.xlsm')[0]
            # print(prefix)

            embed_path = 'xl/embeddings/'

            embedded = []

            count = 0
            with zipfile.ZipFile(filename, 'r') as fd:
                for zipinfo in fd.infolist():
                    fn = zipinfo.filename

                    # Ignore the directory itself, we just want the files.
                    if fn.startswith(embed_path):

                        embedded.append(fn)
                        print("Extracting %s" % fn)

                        fd.extract(zipinfo)

                        # ------------------Store PDF in Database------------------------------
                        str1 = b'%PDF-'  # Begin PDF
                        str2 = b'%%EOF'  # End PDF

                        with open(fn, 'rb') as f:
                            binary_data = f.read()
                        # print(binary_data)

                        # Convert BYTE to BYTEARRAY
                        binary_byte_array = bytearray(binary_data)
                        # print(binary_byte_array)

                        # Find where PDF begins
                        result1 = binary_byte_array.find(str1)
                        # print(result1)

                        # Remove all characters before PDF begins
                        del binary_byte_array[:result1]
                        # print(binary_byte_array)

                        # Find where PDF ends
                        result2 = binary_byte_array.find(str2)
                        # print(result2)

                        # Subtract the length of the array from the position of where PDF ends (add 5 for %%OEF characters)
                        # and delete that many characters from end of array
                        # print(len(binary_byte_array))
                        to_remove = len(binary_byte_array) - (result2 + 5)
                        # print(to_remove)

                        del binary_byte_array[-to_remove:]
                        print(binary_byte_array)

                        #
                        # # Place the byte data into the DB
                        # PDFStorage.create(pdf_byte_data=binary_byte_array)
            #
            #
            # print(embedded)
            #
            # # Retrieve the byte data and create pdf
            # extracted_data = PDFStorage.get(PDFStorage.id == 1).pdf_byte_data
            #
            # with open(os.path.expanduser('test1.pdf'), 'wb') as fout:
            #     fout.write(extracted_data)

            # ------------------Store PDF in Database------------------------------
            # str1 = b'%PDF-'  # Begin PDF
            # str2 = b'%%EOF'  # End PDF
            #
            # with open('oleObject1.bin', 'rb') as f:
            #     binary_data = f.read()
            # print(binary_data)
            #
            # # Convert BYTE to BYTEARRAY
            # binary_byte_array = bytearray(binary_data)
            #
            # # Find where PDF begins
            # result1 = binary_byte_array.find(str1)
            # print(result1)
            #
            # # Remove all characters before PDF begins
            # del binary_byte_array[:result1]
            # print(binary_byte_array)
            #
            # # Find where PDF ends
            # result2 = binary_byte_array.find(str2)
            # print(result2)
            #
            # # Subtract the length of the array from the position of where PDF ends (add 5 for %%OEF characters)
            # # and delete that many characters from end of array
            # print(len(binary_byte_array))
            # to_remove = len(binary_byte_array) - (result2 + 5)
            # print(to_remove)
            #
            # del binary_byte_array[-to_remove:]
            # print(binary_byte_array)
            #
            # # Place the byte data into the DB
            # PDFStorage.create(pdf_byte_data=binary_byte_array)
            #
            # # Retrieve the byte data and create pdf
            # extracted_data = PDFStorage.get(PDFStorage.id == 1).pdf_byte_data
            #
            # with open(os.path.expanduser('test1.pdf'), 'wb') as fout:
            #     fout.write(extracted_data)

            # ------------------Extract Excel Data---------------------------------
            # Sheet 1 G1
            # ws = wbOPENPYXL.get_sheet_by_name('G1 (Personnel Management)')
            #
            # g1_data = {
            #     'date': ws['I3'].value,
            #     'major_command': ws['C5'].value,
            #     'unit': ws['C6'].value,
            #     'installation': ws['C7'].value,
            #     'address': ws['A100'].value,
            #     'city': ws['A101'].value,
            #     'state': ws['A102'].value,
            #     'poc': ws['C9'].value,
            #     'poc_email': ws['C10'].value,
            #     'poc_phone': ws['C11'].value,
            #     'team_member_name': ws['G6'].value,
            #     'team_member_phone': ws['G7'].value,
            #     'team_member_email': ws['G8'].value,
            #     'team_member_signature': ws['G9'].value,
            #     'num_supported_units': ws['C17'].value,
            #     'num_of_personnel': ws['C18'].value,
            #     'union_reps': ws['C20'].value
            # }
            # # Fastest way to INSERT multiple rows.
            # G1.insert_many(g1_data).execute()

            # # Sheet 2 G2
            # sheet = wb.sheet_by_index(2)
            #
            # g2_data = {
            #     'date': sheet.cell_value(2, 6),
            #     'organization': sheet.cell_value(4, 2),
            #     'unit': sheet.cell_value(5, 2),
            #     'address': sheet.cell_value(6, 2),
            #     'point_of_contact': sheet.cell_value(7, 2),
            #     'email': sheet.cell_value(8, 2),
            #     'phone_number': sheet.cell_value(9, 2),
            #     'name': sheet.cell_value(5, 4),
            #     'second_phone_number': sheet.cell_value(6, 4),
            #     'second_email': sheet.cell_value(7, 4),
            #     'signature': sheet.cell_value(8, 4),
            #     'detailed_data_a': sheet.cell_value(13, 2),
            #     'detailed_data_b': sheet.cell_value(15, 2),
            #     'detailed_data_c': sheet.cell_value(17, 2),
            #     'detailed_data_d': sheet.cell_value(19, 2),
            #     'detailed_data_e': sheet.cell_value(21, 2),
            #     'detailed_data_f': sheet.cell_value(23, 2),
            #     'detailed_data_g': sheet.cell_value(25, 2),
            #     'detailed_data_h': sheet.cell_value(27, 2)
            # }
            # G2.insert_many(g2_data).execute()
            #
            # # Sheet 3 G3
            # sheet = wb.sheet_by_index(3)
            #
            # g3_data = {
            #     'date': sheet.cell_value(2, 8),
            #     'organization': sheet.cell_value(4, 2),
            #     'unit': sheet.cell_value(5, 2),
            #     'address': sheet.cell_value(6, 2),
            #     'point_of_contact': sheet.cell_value(7, 2),
            #     'email': sheet.cell_value(8, 2),
            #     'phone_number': sheet.cell_value(9, 2),
            #     'name': sheet.cell_value(5, 6),
            #     'second_phone_number': sheet.cell_value(6, 6),
            #     'second_email': sheet.cell_value(7, 6),
            #     'signature': sheet.cell_value(8, 6),
            #     'detailed_data_a': sheet.cell_value(13, 2),
            #     'detailed_data_b': sheet.cell_value(15, 2),
            #     'detailed_data_c': sheet.cell_value(17, 2),
            #     'detailed_data_d': sheet.cell_value(19, 2),
            #     'detailed_data_e': sheet.cell_value(21, 2),
            #     'detailed_data_f1': sheet.cell_value(23, 2),
            #     'detailed_data_f2': sheet.cell_value(24, 2),
            #     'detailed_data_f3': sheet.cell_value(25, 2),
            #     'detailed_data_f4': sheet.cell_value(26, 2),
            #     'detailed_data_f5': sheet.cell_value(27, 2),
            #     'detailed_data_f6': sheet.cell_value(28, 2),
            # }
            # G3.insert_many(g3_data).execute()
            #
            # # Sheet 4 G4 (Contracts)
            # sheet = wb.sheet_by_index(4)
            #
            # g4c_data = {
            #     'date': sheet.cell_value(2, 8),
            #     'organization': sheet.cell_value(4, 2),
            #     'unit': sheet.cell_value(5, 2),
            #     'address': sheet.cell_value(6, 2),
            #     'point_of_contact': sheet.cell_value(7, 2),
            #     'email': sheet.cell_value(8, 2),
            #     'phone_number': sheet.cell_value(9, 2),
            #     'name': sheet.cell_value(5, 6),
            #     'second_phone_number': sheet.cell_value(6, 6),
            #     'second_email': sheet.cell_value(7, 6),
            #     'signature': sheet.cell_value(8, 6),
            #     'detailed_data_a': sheet.cell_value(13, 2),
            #     'detailed_data_b1': sheet.cell_value(15, 2),
            #     'detailed_data_b2': sheet.cell_value(16, 2),
            #     'detailed_data_b3': sheet.cell_value(17, 2),
            #     'detailed_data_b4': sheet.cell_value(18, 2),
            #     'detailed_data_b5': sheet.cell_value(19, 2),
            #     'detailed_data_b6': sheet.cell_value(20, 2),
            # }
            # G4C.insert_many(g4c_data).execute()
            #
            # # Sheet 5 G4 (Properties and Facilities)
            # sheet = wb.sheet_by_index(5)
            #
            # g4pf_data = {
            #     'date': sheet.cell_value(2, 8),
            #     'organization': sheet.cell_value(4, 2),
            #     'unit': sheet.cell_value(5, 2),
            #     'address': sheet.cell_value(6, 2),
            #     'point_of_contact': sheet.cell_value(7, 2),
            #     'email': sheet.cell_value(8, 2),
            #     'phone_number': sheet.cell_value(9, 2),
            #     'name': sheet.cell_value(5, 6),
            #     'second_phone_number': sheet.cell_value(6, 6),
            #     'second_email': sheet.cell_value(7, 6),
            #     'signature': sheet.cell_value(8, 6),
            #     'detailed_data_a1': sheet.cell_value(13, 2),
            #     'detailed_data_a2': sheet.cell_value(14, 2),
            #     'detailed_data_a3': sheet.cell_value(15, 2),
            #     'detailed_data_a4': sheet.cell_value(16, 2),
            #     'detailed_data_a5': sheet.cell_value(17, 2),
            #     'detailed_data_b1': sheet.cell_value(19, 2),
            #     'detailed_data_b2': sheet.cell_value(20, 2),
            #     'detailed_data_b3': sheet.cell_value(21, 2),
            #     'detailed_data_b4': sheet.cell_value(22, 2),
            # }
            # G4PF.insert_many(g4pf_data).execute()
            #
            # # Sheet 6 G5 Plans
            # sheet = wb.sheet_by_index(6)
            #
            # g5_data = {
            #     'date': sheet.cell_value(2, 8),
            #     'organization': sheet.cell_value(4, 2),
            #     'unit': sheet.cell_value(5, 2),
            #     'address': sheet.cell_value(6, 2),
            #     'point_of_contact': sheet.cell_value(7, 2),
            #     'email': sheet.cell_value(8, 2),
            #     'phone_number': sheet.cell_value(9, 2),
            #     'name': sheet.cell_value(5, 6),
            #     'second_phone_number': sheet.cell_value(6, 6),
            #     'second_email': sheet.cell_value(7, 6),
            #     'signature': sheet.cell_value(8, 6),
            #     'detailed_data_a': sheet.cell_value(13, 2),
            #     'detailed_data_b': sheet.cell_value(15, 2),
            #     'detailed_data_c': sheet.cell_value(17, 2),
            # }
            # G5.insert_many(g5_data).execute()
            #
            # # Sheet 7 G6
            # sheet = wb.sheet_by_index(7)
            #
            # g6_data = {
            #     'date': sheet.cell_value(2, 8),
            #     'organization': sheet.cell_value(4, 2),
            #     'unit': sheet.cell_value(5, 2),
            #     'address': sheet.cell_value(6, 2),
            #     'point_of_contact': sheet.cell_value(7, 2),
            #     'email': sheet.cell_value(8, 2),
            #     'phone_number': sheet.cell_value(9, 2),
            #     'name': sheet.cell_value(5, 6),
            #     'second_phone_number': sheet.cell_value(6, 6),
            #     'second_email': sheet.cell_value(7, 6),
            #     'signature': sheet.cell_value(8, 6),
            #     'detailed_data_a': sheet.cell_value(13, 2),
            #     'detailed_data_b': sheet.cell_value(15, 2),
            #     'detailed_data_c': sheet.cell_value(17, 2),
            # }
            # G6.insert_many(g6_data).execute()
            #
            # # Sheet 8 G8
            # sheet = wb.sheet_by_index(8)
            #
            # g8_data = {
            #     'date': sheet.cell_value(2, 8),
            #     'organization': sheet.cell_value(4, 2),
            #     'unit': sheet.cell_value(5, 2),
            #     'address': sheet.cell_value(6, 2),
            #     'point_of_contact': sheet.cell_value(7, 2),
            #     'email': sheet.cell_value(8, 2),
            #     'phone_number': sheet.cell_value(9, 2),
            #     'name': sheet.cell_value(5, 6),
            #     'second_phone_number': sheet.cell_value(6, 6),
            #     'second_email': sheet.cell_value(7, 6),
            #     'signature': sheet.cell_value(8, 6),
            #     'detailed_data_a1': sheet.cell_value(14, 2),
            #     'detailed_data_a2': sheet.cell_value(15, 2),
            #     'detailed_data_a3': sheet.cell_value(16, 2),
            #     'detailed_data_a4': sheet.cell_value(17, 2),
            #     'detailed_data_b1': sheet.cell_value(20, 2),
            #     'detailed_data_b2': sheet.cell_value(21, 2),
            #     'detailed_data_c1': sheet.cell_value(23, 2),
            #     'detailed_data_c2': sheet.cell_value(24, 2),
            # }
            # G8.insert_many(g8_data).execute()

            # Sheet 9 Admin
            # sheet = wb.sheet_by_index(9)
            #
            # admin_data = {
            #     'admin1': sheet.cell_value(2, 2),
            #     'admin2': sheet.cell_value(3, 2),
            #     'admin3': sheet.cell_value(4, 2),
            #     'admin4': sheet.cell_value(5, 2),
            #     'admin5': sheet.cell_value(6, 2),
            #     'customer1': sheet.cell_value(8, 2),
            #     'customer2': sheet.cell_value(9, 2),
            #     'customer3': sheet.cell_value(10, 2),
            #     'customer4': sheet.cell_value(11, 2),
            #     'customer5': sheet.cell_value(12, 2),
            #     'customer6': sheet.cell_value(13, 2),
            #     'customer7': sheet.cell_value(14, 2),
            #     'name1': sheet.cell_value(16, 3),
            #     'name2': sheet.cell_value(17, 3),
            #     'name3': sheet.cell_value(18, 3),
            #     'name4': sheet.cell_value(19, 3),
            #     'name5': sheet.cell_value(20, 3),
            #     'name6': sheet.cell_value(21, 3),
            #     'phone1': sheet.cell_value(16, 4),
            #     'phone2': sheet.cell_value(17, 4),
            #     'phone3': sheet.cell_value(18, 4),
            #     'phone4': sheet.cell_value(19, 4),
            #     'phone5': sheet.cell_value(20, 4),
            #     'phone6': sheet.cell_value(21, 4),
            #     'email1': sheet.cell_value(16, 5),
            #     'email2': sheet.cell_value(17, 5),
            #     'email3': sheet.cell_value(18, 5),
            #     'email4': sheet.cell_value(19, 5),
            #     'email5': sheet.cell_value(20, 5),
            #     'email6': sheet.cell_value(21, 5),
            #     'tenants1': sheet.cell_value(23, 3),
            #     'tenants2': sheet.cell_value(24, 3),
            #     'tenants3': sheet.cell_value(25, 3),
            #     'tenants4': sheet.cell_value(26, 3),
            #     'tenants5': sheet.cell_value(27, 3),
            #     'tenants6': sheet.cell_value(28, 3),
            #     'tenants7': sheet.cell_value(29, 3),
            # }
            # Admin.insert_many(admin_data).execute()

            # # Sheet 10 Required Docs
            # sheet = wb.sheet_by_index(10)
            # print(sheet.cell_value(3, 3))
            # print(sheet.cell_value(4, 3))
            # print(sheet.cell_value(5, 3))
            # print(sheet.cell_value(6, 3))
            # print(sheet.cell_value(7, 3))
            # print(sheet.cell_value(8, 3))
            # print(sheet.cell_value(9, 3))
            # print(sheet.cell_value(10, 3))
            # print(sheet.cell_value(11, 3))

            # Delete the directory where the embedded objects are stored.
            # try:
            #     shutil.rmtree(embed_path)
            #     os.makedirs(embed_path)
            # except OSError as e:
            #     print("Error: %s - %s." % (e.filename, e.strerror))


            # G1.create(g1_fk_id=1, organization=a)


if __name__ == "__main__":
    main()